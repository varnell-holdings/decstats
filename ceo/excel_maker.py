import tkinter as tk
from tkinter import messagebox
from tkcalendar import DateEntry
import csv
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment


def extract_episodes():
    """Extract episodes for selected date and create Excel file"""
    selected_date = cal.get_date()

    # Convert date to dd-mm-yyyy format to match CSV
    date_str = selected_date.strftime("%d-%m-%Y")

    # Read CSV and filter rows
    matching_rows = []
    csv_file = "episodes.csv"

    if not os.path.exists(csv_file):
        messagebox.showerror("Error", f"File {csv_file} not found!")
        return

    try:
        with open(csv_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row["date"] == date_str:
                    matching_rows.append(row)
    except Exception as e:
        messagebox.showerror("Error", f"Error reading CSV: {str(e)}")
        return

    # Check if any rows found
    if not matching_rows:
        messagebox.showinfo("No Data", f"No episodes found for {date_str}")
        return

    # Create Excel file
    output_filename = f"{selected_date.strftime('%Y-%m-%d')}.xlsx"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Episodes"

        # Write headers
        headers = ["upper", "lower", "anal", "polyp"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header.upper())
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_num, row_data in enumerate(matching_rows, 2):
            ws.cell(row=row_num, column=1, value=row_data["upper"])
            ws.cell(row=row_num, column=2, value=row_data["lower"])
            ws.cell(row=row_num, column=3, value=row_data["anal"])
            ws.cell(row=row_num, column=4, value=row_data["polyp"])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_filename)

        messagebox.showinfo(
            "Success",
            f"Created {output_filename}\n{len(matching_rows)} episodes exported",
        )

        # Open the file
        os.startfile(output_filename)

        # Close the application
        root.destroy()

    except Exception as e:
        messagebox.showerror("Error", f"Error creating Excel file: {str(e)}")


# Create main window
root = tk.Tk()
root.title("Episode Extractor")
root.geometry("350x200")
root.resizable(False, False)

# Create and pack widgets
title_label = tk.Label(
    root, text="Select Date to Extract Episodes", font=("Arial", 12, "bold")
)
title_label.pack(pady=20)

# Date picker
cal = DateEntry(
    root,
    width=20,
    background="darkblue",
    foreground="white",
    borderwidth=2,
    date_pattern="dd-mm-yyyy",
    font=("Arial", 11),
)
cal.pack(pady=10)

# Extract button
extract_btn = tk.Button(
    root,
    text="Extract to Excel",
    command=extract_episodes,
    font=("Arial", 10),
    bg="#4CAF50",
    fg="white",
    padx=20,
    pady=10,
    cursor="hand2",
)
extract_btn.pack(pady=20)

# Center window on screen
root.update_idletasks()
width = root.winfo_width()
height = root.winfo_height()
x = (root.winfo_screenwidth() // 2) - (width // 2)
y = (root.winfo_screenheight() // 2) - (height // 2)
root.geometry(f"{width}x{height}+{x}+{y}")

# Run the application
root.mainloop()
